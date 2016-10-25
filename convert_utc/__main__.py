__author__ = 'weinaguo'
from datetime import datetime
import time
from dateutil import tz
import openpyxl
import pandas as pd


from datetime import tzinfo, timedelta, datetime

ZERO = timedelta(0)
HOUR = timedelta(hours=1)

# A UTC class.

class UTC(tzinfo):
    """UTC"""

    def utcoffset(self, dt):
        return ZERO

    def tzname(self, dt):
        return "UTC"

    def dst(self, dt):
        return ZERO

utc = UTC()

# A class building tzinfo objects for fixed-offset time zones.
# Note that FixedOffset(0, "UTC") is a different way to build a
# UTC tzinfo object.

class FixedOffset(tzinfo):
    """Fixed offset in minutes east from UTC."""

    def __init__(self, offset, name):
        self.__offset = timedelta(minutes = offset)
        self.__name = name

    def utcoffset(self, dt):
        return self.__offset

    def tzname(self, dt):
        return self.__name

    def dst(self, dt):
        return ZERO

# A class capturing the platform's idea of local time.

import time as _time

STDOFFSET = timedelta(seconds = -_time.timezone)
if _time.daylight:
    DSTOFFSET = timedelta(seconds = -_time.altzone)
else:
    DSTOFFSET = STDOFFSET

DSTDIFF = DSTOFFSET - STDOFFSET

class LocalTimezone(tzinfo):

    def utcoffset(self, dt):
        if self._isdst(dt):
            return DSTOFFSET
        else:
            return STDOFFSET

    def dst(self, dt):
        if self._isdst(dt):
            return DSTDIFF
        else:
            return ZERO

    def tzname(self, dt):
        return _time.tzname[self._isdst(dt)]

    def _isdst(self, dt):
        tt = (dt.year, dt.month, dt.day,
              dt.hour, dt.minute, dt.second,
              dt.weekday(), 0, 0)
        stamp = _time.mktime(tt)
        tt = _time.localtime(stamp)
        return tt.tm_isdst > 0

Local = LocalTimezone()


headers = ['user', 'action', 'day', 'activitydate']


def utc2local (utc):
    """
    :type utc: datetime
    :rtype: datetime
    """
    epoch = time.mktime(utc.timetuple())
    offset = datetime.fromtimestamp(epoch) - datetime.utcfromtimestamp(epoch)
    return utc + offset + datetime(utc.year, utc.month, utc.day, tzinfo=Local).dst()



def harmonic_activity_log(wb):
    log = wb.get_sheet_by_name('Harmonic Activity Log')
    email = [log['J' + str(x)].value.split('@')
            if '@' in log['J'+str(x)].value else [log['J'+str(x)].value[:-1], 0]
            for x in range(3, log.max_row+1)]
    user = [x[0]for x in email]
    action = [log['I'+str(x)].value for x in range(3, log.max_row+1)]
    day = [log['G'+str(x)].value for x in range(3, log.max_row+1)]
    activitydate = [log['F'+str(x)].value for x in range(3, log.max_row+1)]
    day_map = {
        0: 'Monday',
        1: 'Tuesday',
        2: 'Wednsday',
        3: 'Thursday',
        4: 'Friday',
        5: 'Satuday',
        6: 'Sunday'
    }

    ret = pd.DataFrame()
    ret['user'] = user
    ret['action'] = action
    ret['activitydate'] = [utc2local(x) for x in activitydate]
    ret['day'] = [day_map[x.weekday()] for x in ret['activitydate']]
    return ret


def main():
    wb = openpyxl.load_workbook('/Users/weinaguo/Desktop/harmonic_activity log_10132016.xlsx')
    result = harmonic_activity_log(wb)
    writer = pd.ExcelWriter('/Users/weinaguo/Desktop/harmonicActivity-out.xlsx')
    result.to_excel(writer)
    writer.save()

if __name__ == '__main__':
    main()

